from __future__ import annotations

import json
import statistics
from pathlib import Path


def _parse_result_dict(result_value: str) -> dict | None:
    """Parse JSON result string into a dictionary."""
    if not isinstance(result_value, str):
        return None
    try:
        parsed = json.loads(result_value)
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, dict) else None


def save_results_to_xlsx(results: list[dict], out_path: Path) -> None:
    """Save experiment results to an Excel file with two sheets.

    Sheet 1: Raw Data - all individual measurements with repeats.
    Sheet 2: Aggregated Data - grouped by scale and sample with mean and std.
    """
    if not results:
        raise ValueError("No results available to save.")

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Install it with: pip install openpyxl")

    # Ensure the file has .xlsx extension
    if out_path.suffix.lower() != '.xlsx':
        out_path = out_path.with_suffix('.xlsx')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # ========== Sheet 1: Raw Data ==========
    ws_raw = wb.create_sheet("Raw Data", 0)
    raw_fieldnames = ["scale", "sample_idx", "repeat", "T", "pH", "F1", "F2", "F3", "status", "result"]
    ws_raw.append(raw_fieldnames)
    for row in results:
        ws_raw.append([row.get(field, "") for field in raw_fieldnames])

    # ========== Sheet 2: Aggregated Data ==========
    ws_agg = wb.create_sheet("Aggregated Data", 1)

    # Group repeated measurements by scale and sample index.
    groups: dict[tuple[str, int], list[dict]] = {}
    for row in results:
        key = (row["scale"], int(row["sample_idx"]))
        groups.setdefault(key, []).append(row)

    # Discover numeric output fields from parsed JSON results.
    output_fields: set[str] = set()
    for row in results:
        parsed = _parse_result_dict(row.get("result", ""))
        if parsed is None:
            continue
        for field, value in parsed.items():
            if isinstance(value, (int, float)):
                output_fields.add(field)

    # Build header for aggregated sheet.
    agg_fieldnames = ["scale", "sample_idx", "T", "pH", "F1", "F2", "F3", "repeats", "ok_count", "error_count", "status"]
    for field in sorted(output_fields):
        agg_fieldnames.append(f"{field}_mean")
        agg_fieldnames.append(f"{field}_std")

    ws_agg.append(agg_fieldnames)

    # Write aggregated rows.
    for (scale, sample_idx), rows in sorted(groups.items()):
        base = rows[0]
        row_out = [
            scale,
            sample_idx,
            base["T"],
            base["pH"],
            base["F1"],
            base["F2"],
            base["F3"],
            len(rows),
        ]

        ok_count = 0
        error_count = 0
        numeric_data: dict[str, list[float]] = {field: [] for field in output_fields}

        for row in rows:
            status = str(row.get("status", ""))
            if status == "ok":
                ok_count += 1
                parsed = _parse_result_dict(row.get("result", ""))
                if parsed is None:
                    continue
                for field in output_fields:
                    value = parsed.get(field)
                    if isinstance(value, (int, float)):
                        numeric_data[field].append(float(value))
            else:
                error_count += 1

        row_out.append(ok_count)
        row_out.append(error_count)
        if error_count == 0:
            row_out.append("ok")
        elif ok_count == 0:
            row_out.append("error")
        else:
            row_out.append("partial")

        for field in sorted(output_fields):
            values = numeric_data[field]
            if values:
                row_out.append(statistics.mean(values))
                row_out.append(statistics.stdev(values) if len(values) > 1 else 0.0)
            else:
                row_out.append("")
                row_out.append("")

        ws_agg.append(row_out)

    # Adjust column widths for readability.
    for ws in [ws_raw, ws_agg]:
        for col_num, col_width in enumerate([12, 12, 10, 8, 8, 8, 8, 8, 12, 20], 1):
            ws.column_dimensions[get_column_letter(col_num)].width = col_width

    wb.save(out_path)


def save_results_to_csv(results: list[dict], out_path: Path) -> None:
    """Wrapper for backward compatibility. Redirects to Excel format."""
    save_results_to_xlsx(results, out_path)
